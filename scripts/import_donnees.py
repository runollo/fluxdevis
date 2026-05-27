"""Import des donnees existantes du tarificateur vers la BDD FluxDevis.

Lit donnees_catalogue.xlsx (offres + options) et contact.xlsx (clients)
depuis le projet tarificateur, et les insere dans PostgreSQL.

Usage:
    cd backend
    python -m scripts.import_donnees
"""

import sys
import os

# Ajouter le backend au path pour les imports app.*
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "backend"))

from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.database import sync_engine, SyncSessionLocal
from app.models.base import Base
from app.models.societe import Societe
from app.models.client import Client
from app.models.offre import Offre
from app.models.option import Option, OptionInclusion

# Chemin vers le projet tarificateur existant
TARIFICATEUR_DIR = Path(__file__).resolve().parent.parent.parent / "tarificateur"
CATALOGUE_PATH = TARIFICATEUR_DIR / "donnees_catalogue.xlsx"
CONTACT_PATH = TARIFICATEUR_DIR / "contact.xlsx"

# Mapping des 10 colonnes d'inclusion (indices 10-19 dans le tuple option)
# vers les noms d'offres dans l'ordre du catalogue
INCLUSION_OFFRE_NAMES = [
    "Presence Web - 5 Pages",
    "Croissance - 8 Pages",
    "Serenite - 15 Pages",
    "Premium - 20 Pages",
    "Shopify Vitrine - Site Catalogue",
    "Shopify Essentiel - 25 references",
    "Shopify Croissance - 100 references",
    "Shopify Serenite - 300 references",
    "Shopify Premium - 1000 references",
    "Shopify Upgrade (V2)",
]


def create_tables():
    """Cree toutes les tables si elles n'existent pas."""
    Base.metadata.create_all(sync_engine)
    print("Tables creees.")


def import_societe(session: Session):
    """Insere la societe emettrice (BLUELINK INNOVATIONS)."""
    existing = session.query(Societe).first()
    if existing:
        print(f"Societe deja presente : {existing.nom}")
        return

    societe = Societe(
        nom="BLUELINK INNOVATIONS",
        forme_juridique="SASU au capital de 1 000 EUR",
        marque="FluXweb",
        adresse="199 rue Helene Boucher",
        cp_ville="34170 Castelnau-le-Lez",
        siret="983 700 519 00017",
        rcs="RCS Montpellier",
        tva_intracom="FR50983700519",
        email="contact@fluxweb.fr",
        site_web="www.fluxweb.fr",
        iban="FR76 3000 3032 5300 0200 7622 234",
        bic="SOGEFRPP",
    )
    session.add(societe)
    session.commit()
    print("Societe BLUELINK INNOVATIONS inseree.")


def import_offres(session: Session) -> dict[str, int]:
    """Importe les offres depuis donnees_catalogue.xlsx. Retourne {nom: id}."""
    if not CATALOGUE_PATH.exists():
        print(f"ATTENTION: {CATALOGUE_PATH} introuvable, import offres ignore.")
        return {}

    wb = load_workbook(CATALOGUE_PATH, read_only=True, data_only=True)
    ws = wb["Offres"]

    offre_map = {}
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        if not row[0]:
            continue

        nom = str(row[0]).strip()
        existing = session.query(Offre).filter_by(nom=nom).first()
        if existing:
            offre_map[nom] = existing.id
            continue

        # Determiner le type_site depuis le nom
        type_site = "Shopify" if "shopify" in nom.lower() else "Webflow"

        tarif_achat = Decimal(str(row[2] or 0))
        taux_marge = Decimal(str(row[3] or 0))
        # Recalculer si la valeur cache Excel est vide
        tarif_vente = row[4]
        if tarif_vente is None or tarif_vente == 0:
            tarif_vente = tarif_achat * (1 + taux_marge)
        tarif_vente = Decimal(str(tarif_vente))

        offre = Offre(
            nom=nom,
            type_site=type_site,
            type_offre=str(row[1] or ""),
            tarif_achat=tarif_achat,
            taux_marge=taux_marge,
            tarif_vente_conseille=tarif_vente,
            pages=int(row[5] or 0),
            heures=int(row[6] or 0),
            commission_apporteur=Decimal(str(row[7] or 0)),
            ordre=idx,
        )
        session.add(offre)
        session.flush()
        offre_map[nom] = offre.id

    session.commit()
    wb.close()
    print(f"{len(offre_map)} offres importees.")
    return offre_map


def import_options(session: Session, offre_map: dict[str, int]):
    """Importe les options depuis donnees_catalogue.xlsx + cree les inclusions."""
    if not CATALOGUE_PATH.exists():
        return

    wb = load_workbook(CATALOGUE_PATH, read_only=True, data_only=True)
    ws = wb["Options"]

    count = 0
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        if not row[0]:
            continue

        code = str(row[0]).strip()
        existing = session.query(Option).filter_by(code=code).first()
        if existing:
            continue

        # Champs source
        heures_setup = Decimal(str(row[21] or 0)) if len(row) > 21 else Decimal("0")
        heures_mensuel = Decimal(str(row[22] or 0)) if len(row) > 22 else Decimal("0")
        prix_heure = Decimal(str(row[23] or 27)) if len(row) > 23 else Decimal("27")
        taux_marge = Decimal(str(row[6] or 0))

        option = Option(
            code=code,
            nom=str(row[1] or ""),
            categorie=str(row[2] or ""),
            type_ligne=str(row[3] or "OPTION_SETUP"),
            heures_setup=heures_setup,
            heures_mensuel=heures_mensuel,
            prix_heure=prix_heure,
            taux_marge=taux_marge,
            prix_hebergement=Decimal(str(row[20] or 0)) if len(row) > 20 else Decimal("0"),
            commentaire=str(row[9]) if len(row) > 9 and row[9] else None,
            selection_regle=str(row[24] or "OPTIONNEL") if len(row) > 24 else "OPTIONNEL",
            quantite_defaut=int(row[25] or 0) if len(row) > 25 else 0,
            unite=str(row[26] or "unite") if len(row) > 26 else "unite",
            ordre=idx,
        )
        option.recalculer_prix()
        session.add(option)
        session.flush()

        # Creer les inclusions (colonnes 10 a 19 = incl_presence ... incl_sh_upgrade)
        for col_idx, offre_name in enumerate(INCLUSION_OFFRE_NAMES):
            flag_idx = 10 + col_idx
            if len(row) > flag_idx and row[flag_idx] and int(row[flag_idx]) == 1:
                offre_id = offre_map.get(offre_name)
                if offre_id:
                    inclusion = OptionInclusion(
                        option_id=option.id,
                        offre_id=offre_id,
                    )
                    session.add(inclusion)

        count += 1

    session.commit()
    wb.close()
    print(f"{count} options importees.")


def import_clients(session: Session):
    """Importe les contacts depuis contact.xlsx."""
    if not CONTACT_PATH.exists():
        print(f"ATTENTION: {CONTACT_PATH} introuvable, import clients ignore.")
        return

    wb = load_workbook(CONTACT_PATH, read_only=True, data_only=True)
    ws = wb["Contact Client"]

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue

        raison_sociale = str(row[0]).strip()
        existing = session.query(Client).filter_by(raison_sociale=raison_sociale).first()
        if existing:
            continue

        client = Client(
            raison_sociale=raison_sociale,
            adresse=str(row[1] or "") if len(row) > 1 and row[1] else None,
            code_postal=str(row[2] or "") if len(row) > 2 and row[2] else None,
            ville=str(row[3] or "") if len(row) > 3 and row[3] else None,
            interlocuteur=str(row[5] or "") if len(row) > 5 and row[5] else None,
            telephone=str(row[8] or "") if len(row) > 8 and row[8] else None,
            siret=str(row[9] or "") if len(row) > 9 and row[9] else None,
        )
        session.add(client)
        count += 1

    session.commit()
    wb.close()
    print(f"{count} clients importes.")


def main():
    print("=" * 60)
    print("IMPORT DONNEES TARIFICATEUR -> FLUXDEVIS")
    print("=" * 60)
    print(f"Catalogue : {CATALOGUE_PATH}")
    print(f"Contacts  : {CONTACT_PATH}")
    print()

    create_tables()

    with SyncSessionLocal() as session:
        import_societe(session)
        offre_map = import_offres(session)
        import_options(session, offre_map)
        import_clients(session)

    print()
    print("Import termine.")


if __name__ == "__main__":
    main()
