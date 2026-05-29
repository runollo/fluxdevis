"""Export Excel (.xlsx) des devis et factures via openpyxl.

Genere un classeur en memoire (BytesIO) pret a etre renvoye en StreamingResponse.
Les colonnes monetaires sont des nombres formates en euros pour rester exploitables
dans un tableur (tris, sommes, filtres).
"""

from decimal import Decimal
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.models.devis import Devis
from app.models.facture import Facture

# Bleu FluXweb pour l'entete
_ENTETE_FILL = PatternFill("solid", fgColor="1A355E")
_ENTETE_FONT = Font(bold=True, color="FFFFFF")
_FORMAT_EUR = '# ##0.00 "EUR"'


def _ecrire_entete(ws, colonnes: list[str]) -> None:
    for i, titre in enumerate(colonnes, start=1):
        cell = ws.cell(row=1, column=i, value=titre)
        cell.fill = _ENTETE_FILL
        cell.font = _ENTETE_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def _ajuster_largeurs(ws, largeurs: list[int]) -> None:
    for i, largeur in enumerate(largeurs, start=1):
        ws.column_dimensions[get_column_letter(i)].width = largeur


def _num(v) -> float:
    return float(v) if v is not None else 0.0


def export_devis_xlsx(devis: Iterable[Devis]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Devis"
    colonnes = [
        "Reference", "Date", "Client", "Offre", "Mode",
        "Plan", "Statut", "Total HT", "TVA", "Total TTC",
    ]
    _ecrire_entete(ws, colonnes)

    for d in devis:
        ws.append([
            d.reference,
            d.date_emission.strftime("%d/%m/%Y") if d.date_emission else "",
            d.client_raison_sociale,
            d.offre_nom,
            d.mode_reglement.value if d.mode_reglement else "",
            d.plan_paiement.value if d.plan_paiement else "",
            d.statut.value if d.statut else "",
            _num(d.total_ht),
            _num(d.total_tva),
            _num(d.total_ttc),
        ])

    for col in (8, 9, 10):  # colonnes monetaires
        for cell in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            cell[0].number_format = _FORMAT_EUR

    _ajuster_largeurs(ws, [22, 12, 28, 26, 10, 12, 12, 14, 12, 14])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_factures_xlsx(factures: Iterable[Facture]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Factures"
    colonnes = [
        "Numero", "Type", "Statut", "Emission", "Echeance",
        "Objet", "Total HT", "TVA", "Total TTC", "Paiement",
    ]
    _ecrire_entete(ws, colonnes)

    for f in factures:
        ws.append([
            f.numero,
            f.type.value if f.type else "",
            f.statut.value if f.statut else "",
            f.date_emission.strftime("%d/%m/%Y") if f.date_emission else "",
            f.date_echeance.strftime("%d/%m/%Y") if f.date_echeance else "",
            f.objet,
            _num(f.total_ht),
            _num(f.total_tva),
            _num(f.total_ttc),
            f.date_paiement.strftime("%d/%m/%Y") if f.date_paiement else "",
        ])

    for col in (7, 8, 9):  # colonnes monetaires
        for cell in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            cell[0].number_format = _FORMAT_EUR

    _ajuster_largeurs(ws, [22, 14, 12, 12, 12, 40, 14, 12, 14, 12])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
